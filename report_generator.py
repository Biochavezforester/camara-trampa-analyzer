"""
Generador de reportes y exportación dual de Excel.
Genera Excel básico (FORXIME/2) y completo (con trazabilidad IA).
"""

import pandas as pd
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


class ExcelExporter:
    """Exportador de datos a Excel con formato dual."""
    
    @staticmethod
    def export_basic_excel(df: pd.DataFrame, output_path: Path, 
                          coordinates_df: Optional[pd.DataFrame] = None) -> Path:
        """
        Exporta Excel básico compatible con FORXIME/2.
        
        Solo columnas: SITIO, CAMARA, ESPECIE, FECHA, HORA
        
        Args:
            df: DataFrame con datos validados
            output_path: Ruta de salida
            coordinates_df: DataFrame opcional con coordenadas
            
        Returns:
            Path al archivo generado
        """
        # Seleccionar solo columnas obligatorias para FORXIME/2
        basic_columns = ['SITIO', 'CAMARA', 'ESPECIE', 'FECHA', 'HORA']
        
        # Asegurar que existan todas las columnas
        for col in basic_columns:
            if col not in df.columns:
                raise ValueError(f"Columna requerida '{col}' no encontrada en DataFrame")
        
        df_basic = df[basic_columns].copy()
        
        # Limpiar datos
        df_basic = ExcelExporter._clean_dataframe(df_basic)
        
        # Crear archivo Excel
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Hoja principal de datos
            df_basic.to_excel(writer, sheet_name='Datos', index=False)
            
            # Hoja de coordenadas si está disponible
            if coordinates_df is not None and len(coordinates_df) > 0:
                coordinates_df.to_excel(writer, sheet_name='Coordenadas', index=False)
            
            # Formatear hojas
            ExcelExporter._format_worksheet(writer.sheets['Datos'], df_basic)
            if 'Coordenadas' in writer.sheets:
                ExcelExporter._format_worksheet(writer.sheets['Coordenadas'], coordinates_df)
        
        return output_path
    
    @staticmethod
    def export_complete_excel(df: pd.DataFrame, output_path: Path,
                             effort_df: Optional[pd.DataFrame] = None,
                             events_df: Optional[pd.DataFrame] = None,
                             temporal_df: Optional[pd.DataFrame] = None,
                             coordinates_df: Optional[pd.DataFrame] = None,
                             summary: Optional[Dict] = None) -> Path:
        """
        Exporta Excel completo con todas las columnas y análisis.
        
        Incluye: datos completos, coordenadas, esfuerzo, eventos independientes,
        análisis temporal, resumen ejecutivo.
        
        Args:
            df: DataFrame con datos completos (incluye predicciones IA si existen)
            output_path: Ruta de salida
            effort_df: DataFrame de esfuerzo de muestreo
            events_df: DataFrame de eventos independientes
            temporal_df: DataFrame de análisis temporal
            coordinates_df: DataFrame de coordenadas
            summary: Dict con resumen ejecutivo
            
        Returns:
            Path al archivo generado
        """
        # Limpiar datos
        df_clean = ExcelExporter._clean_dataframe(df)
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Hoja 1: Datos completos
            df_clean.to_excel(writer, sheet_name='Datos', index=False)
            ExcelExporter._format_worksheet(writer.sheets['Datos'], df_clean)
            
            # Hoja 2: Coordenadas
            if coordinates_df is not None and len(coordinates_df) > 0:
                coordinates_df.to_excel(writer, sheet_name='Coordenadas', index=False)
                ExcelExporter._format_worksheet(writer.sheets['Coordenadas'], coordinates_df)
            
            # Hoja 3: Esfuerzo de muestreo
            if effort_df is not None and len(effort_df) > 0:
                effort_df.to_excel(writer, sheet_name='Esfuerzo', index=False)
                ExcelExporter._format_worksheet(writer.sheets['Esfuerzo'], effort_df)
            
            # Hoja 4: Eventos independientes
            if events_df is not None and len(events_df) > 0:
                events_df.to_excel(writer, sheet_name='Eventos_Independientes', index=False)
                ExcelExporter._format_worksheet(writer.sheets['Eventos_Independientes'], events_df)
            
            # Hoja 5: Análisis temporal
            if temporal_df is not None and len(temporal_df) > 0:
                temporal_df.to_excel(writer, sheet_name='Analisis_Temporal', index=False)
                ExcelExporter._format_worksheet(writer.sheets['Analisis_Temporal'], temporal_df)
            
            # Hoja 6: Resumen ejecutivo
            if summary:
                ExcelExporter._create_summary_sheet(writer, summary)
        
        return output_path
    
    @staticmethod
    def _clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
        """Limpia DataFrame para exportación."""
        df_clean = df.copy()
        
        # Eliminar espacios en blanco al inicio/final
        for col in df_clean.select_dtypes(include=['object']).columns:
            df_clean[col] = df_clean[col].astype(str).str.strip()
        
        # Reemplazar NaN con cadena vacía
        df_clean = df_clean.fillna('')
        
        return df_clean
    
    @staticmethod
    def _format_worksheet(ws, df: pd.DataFrame):
        """Aplica formato a hoja de Excel."""
        # Formato de encabezados
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Congelar primera fila
        ws.freeze_panes = 'A2'
    
    @staticmethod
    def _create_summary_sheet(writer, summary: Dict):
        """Crea hoja de resumen ejecutivo."""
        wb = writer.book
        ws = wb.create_sheet('Resumen')
        
        # Título
        ws['A1'] = 'RESUMEN EJECUTIVO'
        ws['A1'].font = Font(size=16, bold=True, color="2E7D32")
        
        row = 3
        
        # Información general
        ws[f'A{row}'] = 'INFORMACIÓN GENERAL'
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1
        
        general_info = [
            ('Total de sitios:', summary.get('total_sites', 0)),
            ('Total de cámaras:', summary.get('total_cameras', 0)),
            ('Total de especies detectadas:', summary.get('total_species', 0)),
            ('Total de capturas:', summary.get('total_captures', 0)),
            ('Esfuerzo total (trampas-día):', summary.get('total_trap_days', 0)),
        ]
        
        for label, value in general_info:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1
        
        row += 1
        
        # Información de IA si existe
        if 'ai_predictions' in summary:
            ws[f'A{row}'] = 'CLASIFICACIÓN CON IA'
            ws[f'A{row}'].font = Font(size=12, bold=True)
            row += 1
            
            ai_info = [
                ('Total de predicciones IA:', summary.get('ai_predictions', 0)),
                ('Predicciones validadas:', summary.get('validated_predictions', 0)),
                ('Predicciones corregidas:', summary.get('corrected_predictions', 0)),
                ('Precisión de IA:', f"{summary.get('ai_accuracy', 0):.1f}%"),
            ]
            
            for label, value in ai_info:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                row += 1
            
            row += 1
        
        # Especies más frecuentes
        if 'top_species' in summary:
            ws[f'A{row}'] = 'ESPECIES MÁS FRECUENTES'
            ws[f'A{row}'].font = Font(size=12, bold=True)
            row += 1
            
            ws[f'A{row}'] = 'Especie'
            ws[f'B{row}'] = 'Capturas'
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'].font = Font(bold=True)
            row += 1
            
            for species, count in summary['top_species']:
                ws[f'A{row}'] = species
                ws[f'B{row}'] = count
                row += 1
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20


class ExecutiveSummaryGenerator:
    """Generador de resumen ejecutivo."""
    
    @staticmethod
    def generate_summary(df: pd.DataFrame, effort_df: Optional[pd.DataFrame] = None,
                        events_df: Optional[pd.DataFrame] = None,
                        ai_stats: Optional[Dict] = None) -> Dict:
        """
        Genera resumen ejecutivo del proyecto.
        
        Args:
            df: DataFrame con datos
            effort_df: DataFrame de esfuerzo
            events_df: DataFrame de eventos independientes
            ai_stats: Estadísticas de IA si existen
            
        Returns:
            Dict con resumen
        """
        summary = {
            'total_sites': df['SITIO'].nunique(),
            'total_cameras': df['CAMARA'].nunique(),
            'total_species': df['ESPECIE'].nunique(),
            'total_captures': len(df),
            'total_trap_days': effort_df['TRAMPAS_DIA'].sum() if effort_df is not None else 0,
            'top_species': df['ESPECIE'].value_counts().head(10).items(),
            'date_range': {
                'start': df['FECHA'].min(),
                'end': df['FECHA'].max()
            }
        }
        
        # Agregar estadísticas de eventos independientes si existen
        if events_df is not None:
            summary['total_independent_events'] = events_df['EVENTOS_INDEPENDIENTES'].sum()
        
        # Agregar estadísticas de IA si existen
        if ai_stats:
            summary.update(ai_stats)
        
        return summary


def export_dual_excel(df: pd.DataFrame, project_path: Path, project_name: str,
                     effort_df: Optional[pd.DataFrame] = None,
                     events_df: Optional[pd.DataFrame] = None,
                     temporal_df: Optional[pd.DataFrame] = None,
                     coordinates_df: Optional[pd.DataFrame] = None,
                     ai_stats: Optional[Dict] = None) -> Tuple[Path, Path]:
    """
    Exporta ambos archivos Excel: básico y completo.
    
    Returns:
        Tupla (path_basic, path_complete)
    """
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    # Generar resumen
    summary = ExecutiveSummaryGenerator.generate_summary(df, effort_df, events_df, ai_stats)
    
    # Archivo básico (FORXIME/2)
    basic_filename = f"{project_name}_FORXIME2_{timestamp}.xlsx"
    basic_path = project_path / basic_filename
    ExcelExporter.export_basic_excel(df, basic_path, coordinates_df)
    
    # Archivo completo
    complete_filename = f"{project_name}_COMPLETO_{timestamp}.xlsx"
    complete_path = project_path / complete_filename
    ExcelExporter.export_complete_excel(
        df, complete_path, effort_df, events_df, 
        temporal_df, coordinates_df, summary
    )
    
    return basic_path, complete_path
